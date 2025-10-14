from flask import Flask, render_template, request, jsonify, redirect, url_for
import openpyxl
import os

app = Flask(__name__)

EXCEL_PATH = 'inventory.xlsx'

def get_inventory_headers(excel_path=EXCEL_PATH):
    if not os.path.exists(excel_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        default_headers = ['Barcode', 'Product Name', 'Quantity', 'Price']
        ws.append(default_headers)
        wb.save(excel_path)
        return default_headers
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    return headers

def find_product_by_barcode(barcode, excel_path=EXCEL_PATH):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    barcode_column = None
    for idx, header in enumerate(headers):
        if str(header).lower() == "barcode":
            barcode_column = idx
            break
    if barcode_column is None:
        return None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[barcode_column]).strip() == str(barcode).strip():
            return dict(zip(headers, row))
    return None

@app.route('/scan')
def scan():
    return render_template('index.html')

@app.route('/save_barcode', methods=['POST'])
def save_barcode():
    data = request.get_json()
    barcode = data.get('barcode')
    product = find_product_by_barcode(barcode)
    if product:
        return jsonify({"fields": product})
    else:
        return jsonify({"error": "Barcode not found in inventory."})

# Updated route: Guide the user to use the Streamlit app for adding products
@app.route('/add_product_page', methods=['GET'])
def add_product_page():
    # Instead of rendering add_product.html, show a message with a link to Streamlit
    # Assuming your Streamlit app runs on port 8501
    streamlit_url = "http://localhost:8501"
    return f"""
    <h2>Add Product</h2>
    <p>The Add Product page is now managed by Streamlit.</p>
    <p>
        <a href="{streamlit_url}" target="_blank">Open Add Product Page</a>
    </p>
    <p>Or run <code>streamlit run Inventory_Manager.py</code> in your terminal.</p>
    """

if __name__ == '__main__':
    app.run(port=5001)